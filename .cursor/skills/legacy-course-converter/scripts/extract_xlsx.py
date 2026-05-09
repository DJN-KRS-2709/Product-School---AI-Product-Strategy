#!/usr/bin/env python3
"""Extract an .xlsx (or .csv) file to markdown tables in insights/extracted/.

Usage:
    python extract_xlsx.py <path-to-file.xlsx> [--out <output-dir>] [--include-hidden]

Each sheet becomes a `## Sheet: <name>` section with a markdown table.
Merged cells are unmerged with the value back-filled into each cell.
Cached formula values are exported (re-open in Excel once if values are None).
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "openpyxl not installed. Run: pip install -r requirements.txt\n"
    )
    sys.exit(1)


def render_csv(in_path: Path) -> str:
    with in_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        return "_(empty file)_"
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    out: list[str] = []
    out.append("| " + " | ".join(c.strip() or " " for c in rows[0]) + " |")
    out.append("| " + " | ".join(["---"] * width) + " |")
    for r in rows[1:]:
        out.append("| " + " | ".join((c.strip() or " ").replace("\n", " ") for c in r) + " |")
    return "\n".join(out)


def unmerge_with_fill(sheet) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)
    for rng in merged_ranges:
        top_left_value = sheet.cell(row=rng.min_row, column=rng.min_col).value
        sheet.unmerge_cells(str(rng))
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                sheet.cell(row=r, column=c, value=top_left_value)


def render_sheet(sheet) -> str:
    if sheet.max_row is None or sheet.max_row == 0:
        return "_(empty sheet)_"

    unmerge_with_fill(sheet)

    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        cells = [
            ("" if v is None else str(v)).strip().replace("\n", " ")
            for v in row
        ]
        if any(c for c in cells):
            rows.append(cells)

    if not rows:
        return "_(empty sheet)_"

    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]

    out: list[str] = []
    out.append("| " + " | ".join(c or " " for c in rows[0]) + " |")
    out.append("| " + " | ".join(["---"] * width) + " |")
    for r in rows[1:]:
        out.append("| " + " | ".join(c or " " for c in r) + " |")
    return "\n".join(out)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="Path to the .xlsx or .csv file")
    parser.add_argument("--out", default=None, help="Output directory")
    parser.add_argument(
        "--include-hidden", action="store_true", help="Include hidden sheets"
    )
    args = parser.parse_args()

    in_path = Path(args.input).expanduser().resolve()
    if not in_path.exists():
        sys.stderr.write(f"File not found: {in_path}\n")
        return 1

    out_dir = (
        Path(args.out).expanduser().resolve()
        if args.out
        else Path.cwd() / "insights" / "extracted"
    )
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{in_path.stem}.md"

    blocks: list[str] = [
        f"# {in_path.name}",
        "",
        f"Extracted from `{in_path}`.",
        "",
        "---",
        "",
    ]

    if in_path.suffix.lower() == ".csv":
        blocks.append(f"## Sheet: {in_path.stem}")
        blocks.append("")
        blocks.append(render_csv(in_path))
        blocks.append("")
    else:
        wb = load_workbook(str(in_path), data_only=True, read_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.sheet_state != "visible" and not args.include_hidden:
                blocks.append(f"## Sheet: {sheet_name}  _(hidden — skipped)_")
                blocks.append("")
                continue
            blocks.append(f"## Sheet: {sheet_name}")
            blocks.append("")
            blocks.append(render_sheet(ws))
            blocks.append("")

    out_path.write_text("\n".join(blocks).rstrip() + "\n", encoding="utf-8")
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
